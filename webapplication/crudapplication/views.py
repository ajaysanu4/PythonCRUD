
from crudapplication.forms import EmployeeForm

import openpyxl

from django.shortcuts import render, redirect

import pandas

import csv
from django.http import HttpResponse
from crudapplication.models import Employee

from django.shortcuts import render
from crudapplication.forms import UserForm,UserProfileInfoForm
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required



# Create your views here.
def emp(request):
    if request.method == "POST":
        form = EmployeeForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect("/show")
            except:
                pass

    else:
        form = EmployeeForm()
    return render(request, "Index.html", {'form': form})


def show(request):
    employees = Employee.objects.all()
    return render(request, "show.html", {'employees': employees})


def edit(request, id):
    employee = Employee.objects.get(id=id)
    return render(request, "edit.html", {'employee': employee})


def update(request, id):
    employee = Employee.objects.get(id=id)
    form = EmployeeForm(request.POST, instance=employee)
    if form.is_valid():
        form.save()
        return redirect("/show")
    return render(request, "edit.html", {'employee': employee})


def delete(request, id):
    employee = Employee.objects.get(id=id)
    employee.delete()
    return redirect("/show")




def export_users_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    writer = csv.writer(response)
    writer.writerow(['Employee Id', 'Employee Name', 'Employee Email', 'Employee Contact'])


    list1=['eid','ename']
    string=(", ".join(repr(e) for e in list1).replace(""","").replace(""", ""))
    print (string)
    for data in list1:
        list = Employee.objects.all().values_list(string).replace('\"', "")
        print(list)
    for element in list:
          writer.writerow(element)


    return response

def index(request):
    if "GET" == request.method:
        return render(request, 'show.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment';
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets

        sheet = wb.get_sheet_by_name('Sheet1')

        lists=[]
        employees = Employee.objects.all()
        for employee in employees:
            lists.append(employee)
         #   print(employee)


        for list in lists:
            print(list.eid)
            print(list.ename)

            if sheet.cell(row=2, column=1)=="Employee Id":
                 c1 = sheet.cell(row=2, column=1)
                 c1.value = list.eid
                 stock=pandas.read_csv('C:/Users/Arjun/Desktop/New XLSX Worksheet.xslx')
                 lastprice = stock.iloc[1,3]
                 print(lastprice)
                 lastprice.value=list.ename
           # for i in range(len(lists)):
                 # sheet['A' + str(i)] =list.eid
                 # sheet['B' + str(i)] =list.ename
                 # sheet['C' + str(i)] = list.eemail
                 # sheet['D' + str(i)] = list.econtact

        wb.save(excel_file)

        return redirect("/show")



def index2(request):
    return render(request,'index2.html')
@login_required
def special(request):
    return HttpResponse("You are logged in !")
@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect(reverse('index2'))
def register(request):
    registered = False
    if request.method == 'POST':
        user_form = UserForm(data=request.POST)
        profile_form = UserProfileInfoForm(data=request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            user.set_password(user.password)
            user.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            if 'profile_pic' in request.FILES:
                print('found it')
                profile.profile_pic = request.FILES['profile_pic']
            profile.save()
            registered = True
        else:
            print(user_form.errors,profile_form.errors)
    else:
        user_form = UserForm()
        profile_form = UserProfileInfoForm()
    return render(request,'register.html',
                          {'user_form':user_form,
                           'profile_form':profile_form,
                           'registered':registered})

def user_login(request):
    print("*****")
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            if user.is_active:
                login(request,user)
                print("*****1")
                return HttpResponseRedirect(reverse('show'))
            else:
                return HttpResponse("Your account was inactive.")
        else:
            print("*****2")
            print("Someone tried to login and failed.")
            print("They used username: {} and password: {}".format(username,password))
            return HttpResponse("Invalid login details given")
    else:
        return render(request, 'login.html', {})

